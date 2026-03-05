"""AI投稿企画 - ユーザーのアイデアからバズる構成案を自動生成"""
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

import pandas as pd

import sys
sys.path.insert(0, str(Path(__file__).parent))
from config import DATA_DIR, CATEGORIES, GEMINI_API_KEY


def _find_similar_posts(idea: str, top_n: int = 10) -> Dict[str, Any]:
    """ユーザーのアイデアに類似するバズ投稿を検索"""
    results = {"own_posts": [], "trend_posts": [], "analysis_summary": {}}

    # 自分の投稿データから類似投稿を検索
    ig_path = DATA_DIR / "instagram_analyzed.csv"
    if ig_path.exists():
        ig_df = pd.read_csv(ig_path)
        idea_lower = idea.lower()

        # キーワードマッチでスコアリング
        keywords = re.findall(r"[a-zA-Zぁ-んァ-ヶ亜-熙\w]+", idea_lower)
        if "caption" in ig_df.columns:
            ig_df["_match_score"] = ig_df["caption"].fillna("").apply(
                lambda cap: sum(1 for kw in keywords if kw in cap.lower())
            )

            # カテゴリ分類でもマッチ
            for cat, kws in CATEGORIES.items():
                if any(kw in idea_lower for kw in kws):
                    if "primary_category" in ig_df.columns:
                        ig_df.loc[ig_df["primary_category"] == cat, "_match_score"] += 2

            # エンゲージメント率でソート（マッチ投稿優先）
            er_col = "engagement_rate" if "engagement_rate" in ig_df.columns else None
            if er_col:
                matched = ig_df[ig_df["_match_score"] > 0].sort_values(
                    [er_col], ascending=False
                ).head(top_n)

                if len(matched) < top_n:
                    top_all = ig_df.sort_values(er_col, ascending=False).head(top_n)
                    matched = pd.concat([matched, top_all]).drop_duplicates(subset=["id"]).head(top_n)

                for _, row in matched.iterrows():
                    results["own_posts"].append({
                        "caption": str(row.get("caption", ""))[:300],
                        "media_type": row.get("media_type", ""),
                        "engagement_rate": round(row.get("engagement_rate", 0), 2),
                        "reach": int(row.get("reach", 0)),
                        "saved": int(row.get("saved", 0)),
                        "like_count": int(row.get("like_count", 0)),
                        "category": row.get("primary_category", ""),
                    })

        # 分析サマリー
        if er_col and not ig_df.empty:
            results["analysis_summary"]["avg_engagement_rate"] = round(ig_df[er_col].mean(), 2)
            results["analysis_summary"]["best_media_type"] = ig_df.groupby("media_type")[er_col].mean().idxmax()
            if "primary_category" in ig_df.columns:
                cat_perf = ig_df.groupby("primary_category")[er_col].mean().sort_values(ascending=False)
                results["analysis_summary"]["category_ranking"] = cat_perf.to_dict()
            if "hour" in ig_df.columns:
                hour_perf = ig_df.groupby("hour")[er_col].mean().sort_values(ascending=False)
                results["analysis_summary"]["best_hours"] = hour_perf.head(3).index.tolist()
            if "day_of_week_jp" in ig_df.columns:
                day_perf = ig_df.groupby("day_of_week_jp")[er_col].mean().sort_values(ascending=False)
                results["analysis_summary"]["best_days"] = day_perf.head(3).index.tolist()

    # トレンドリサーチデータから類似投稿を検索
    research_path = DATA_DIR / "hashtag_research_latest.csv"
    if research_path.exists():
        research_df = pd.read_csv(research_path)
        if "caption" in research_df.columns and "engagement" in research_df.columns:
            research_df["_match_score"] = research_df["caption"].fillna("").apply(
                lambda cap: sum(1 for kw in keywords if kw in str(cap).lower())
            )

            top_research = research_df.sort_values(
                ["_match_score", "engagement"], ascending=[False, False]
            ).head(top_n)

            for _, row in top_research.iterrows():
                results["trend_posts"].append({
                    "caption": str(row.get("caption", ""))[:300],
                    "media_type": row.get("media_type", ""),
                    "engagement": int(row.get("engagement", 0)),
                    "like_count": int(row.get("like_count", 0)),
                    "hashtag": row.get("hashtag", ""),
                    "category": row.get("primary_category", ""),
                })

    return results


def _build_prompt(idea: str, similar_data: Dict[str, Any]) -> str:
    """AI用のプロンプトを構築"""

    own_posts_text = ""
    if similar_data["own_posts"]:
        own_posts_text = "### 自分の過去投稿（パフォーマンスが高い順）\n"
        for i, p in enumerate(similar_data["own_posts"][:5], 1):
            own_posts_text += (
                f"{i}. [ER:{p['engagement_rate']}% / リーチ:{p['reach']} / "
                f"保存:{p['saved']} / いいね:{p['like_count']}]\n"
                f"   タイプ: {p['media_type']} / カテゴリ: {p['category']}\n"
                f"   キャプション: {p['caption'][:150]}\n\n"
            )

    trend_posts_text = ""
    if similar_data["trend_posts"]:
        trend_posts_text = "### トレンドリサーチ（高エンゲージメント投稿）\n"
        for i, p in enumerate(similar_data["trend_posts"][:5], 1):
            trend_posts_text += (
                f"{i}. [エンゲージメント:{p['engagement']} / いいね:{p['like_count']}]\n"
                f"   タイプ: {p['media_type']} / ハッシュタグ: #{p['hashtag']}\n"
                f"   キャプション: {p['caption'][:150]}\n\n"
            )

    analysis_text = ""
    summary = similar_data.get("analysis_summary", {})
    if summary:
        analysis_text = "### アカウント分析サマリー\n"
        if "avg_engagement_rate" in summary:
            analysis_text += f"- 平均ER: {summary['avg_engagement_rate']}%\n"
        if "best_media_type" in summary:
            analysis_text += f"- 最も反応の良い投稿タイプ: {summary['best_media_type']}\n"
        if "category_ranking" in summary:
            cats = list(summary["category_ranking"].items())[:3]
            analysis_text += f"- カテゴリ別ER: {', '.join(f'{c}={v:.1f}%' for c, v in cats)}\n"
        if "best_hours" in summary:
            analysis_text += f"- 最適投稿時間: {summary['best_hours']}時\n"
        if "best_days" in summary:
            analysis_text += f"- 最適投稿曜日: {summary['best_days']}\n"
        analysis_text += "\n"

    # 戦略アイデンティティの読み込み
    strategy_context = ""
    strategy_path = DATA_DIR / "strategy_identity.json"
    if strategy_path.exists():
        with open(strategy_path, "r", encoding="utf-8") as f:
            strategy = json.load(f)

        pillars = strategy.get("content_strategy_pillars", [])
        pillars_text = "\n".join(
            f"  {i+1}. {p['pillar']}: {p['description']}（原則: {p['key_principle']}）"
            for i, p in enumerate(pillars)
        )

        audiences = strategy.get("target_audiences", [])
        audience_text = "\n".join(
            f"  - {a['segment']}: {a['needs']}" for a in audiences
        )

        traps = strategy.get("critical_analysis", {})
        traps_text = "\n".join(
            f"  - {v['problem']} → {v['solution']}"
            for v in traps.values()
        )

        checklist = strategy.get("pre_posting_checklist", [])
        checklist_text = "\n".join(
            f"  {i+1}. {c['question']}（例: {c['example']}）"
            for i, c in enumerate(checklist)
        )

        coach_directives = "\n".join(
            f"  - {d}" for d in strategy.get("coach_directive", [])
        )

        strategy_context = f"""
## Kouのアイデンティティ定義
- Core Identity: {strategy['identity']['core_identity']}
- Mission: {strategy['identity']['mission']}
- Values: {', '.join(strategy['identity']['values'])}

## コンテンツ戦略の3本柱
{pillars_text}

## ターゲットオーディエンスと心理ニーズ
{audience_text}

## 回避すべき罠
{traps_text}

## 構成案に必ず反映すべき投稿前チェック
{checklist_text}

## コーチとしての指針
{coach_directives}
"""

    prompt = f"""あなたはKouの専属コンテンツプロデューサー兼ビジネスコーチだ。
忖度なしの批判的・客観的視点で、ユーザーのアイデアを「バズるだけでなく、ポジティブな影響を残す投稿」に昇華する構成案を作成せよ。

「完成された絵画」ではなく「動いている物語」を作れ。
冒頭1秒で視聴者の感情を奪い、Kouの「退屈への拒絶」という哲学を1滴混ぜろ。
{strategy_context}
## ユーザー情報
- アカウント: @k_slf_imp (Kou)
- ジャンル: 暮らし・旅行・大学生の日常（筋トレ・美容等の自分磨き）
- フォロワー: 約14,000人
- ターゲット: 同世代の大学生・高校生（男女）、20〜30代女性

## コンテンツアイデア
{idea}

## 参考データ
{analysis_text}
{own_posts_text}
{trend_posts_text}

## 出力フォーマット
以下のJSON形式で構成案を出力してください。必ず有効なJSONとして出力してください。

```json
{{
  "title": "コンテンツタイトル",
  "format": "リール or カルーセル",
  "duration": "動画尺（例: 30〜45秒）",
  "audio": "音声タイプ（例: 本人音声 + BGM）",
  "outfit_location": "衣装・ロケーションの概要",
  "music": "音源の提案（例: トレンド音源 / 落ち着いたBGM）",
  "caption": "キャプション全文（ハッシュタグ含む）",
  "hook": "冒頭フック文（最初の1行で興味を引く文）",
  "posting_time": "推奨投稿時間（例: 金曜日 20:00）",
  "scenes": [
    {{
      "scene_number": 1,
      "duration": "2秒",
      "scene_description": "シーンの説明",
      "location": "ロケーション",
      "outfit": "想定着用衣装",
      "time_of_day": "想定時刻",
      "video_action": "映像・動作の詳細な説明",
      "text_overlay": "想定テロップ（文字入れ）",
      "narration": "想定セリフ（音声）"
    }}
  ],
  "tips": "この投稿を成功させるためのポイント・注意点",
  "reference_analysis": "参考データから得た知見と、この企画に活かしたポイント"
}}
```

## 注意事項
- シーンは5〜8個程度で構成
- 冒頭1〜2秒で視聴者を惹きつけるフックを必ず入れる（「誰の、どの感情を奪うか」を明確に）
- 各シーンの映像・動作は具体的に（カメラワーク、演出、表情なども含めて）
- テロップは文字入れ想定なので、実際にテロップとして表示するテキストを記載
- セリフは実際に話す内容をそのまま記載
- ハッシュタグは10〜15個（ジャンル関連 + トレンド + ニッチ）
- 参考データの分析結果をもとに、投稿タイプ・時間帯・カテゴリを最適化
- 「早起きしました」という結果ではなく、「なぜ」という哲学をキャプションやナレーションに必ず含める
- 3本柱（憧れ / 知的な集合知 / 等身大の人間味）のどれを主軸にするか意識する
- 「完成された絵画」ではなく「動いている物語」——プロセスや葛藤を含める
- JSONのみを出力し、それ以外のテキストは含めないでください
"""
    return prompt


def generate_ai_plan(idea: str) -> Optional[Dict[str, Any]]:
    """AIでコンテンツ構成案を生成"""
    if not GEMINI_API_KEY:
        print("Error: GEMINI_API_KEY が設定されていません。.env に追加してください。")
        return None

    from google import genai

    # 類似投稿を検索
    print("類似投稿を検索中...")
    similar_data = _find_similar_posts(idea)
    print(f"  自分の投稿: {len(similar_data['own_posts'])}件")
    print(f"  トレンド投稿: {len(similar_data['trend_posts'])}件")

    # プロンプト構築
    prompt = _build_prompt(idea, similar_data)

    # Gemini API 呼び出し
    print("AI構成案を生成中...")
    client = genai.Client(api_key=GEMINI_API_KEY)

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
        config={
            "temperature": 0.8,
            "max_output_tokens": 8192,
        },
    )

    response_text = response.text.strip()

    # JSON抽出（```json ... ``` で囲まれていた場合にも対応）
    json_match = re.search(r"```json\s*(.*?)\s*```", response_text, re.DOTALL)
    if json_match:
        response_text = json_match.group(1)

    try:
        plan = json.loads(response_text)
    except json.JSONDecodeError:
        # JSON部分だけ抽出を試みる
        brace_start = response_text.find("{")
        brace_end = response_text.rfind("}") + 1
        if brace_start >= 0 and brace_end > brace_start:
            try:
                plan = json.loads(response_text[brace_start:brace_end])
            except json.JSONDecodeError:
                print(f"JSONパースエラー: {response_text[:200]}")
                return None
        else:
            print(f"JSON形式の応答が得られませんでした: {response_text[:200]}")
            return None

    # メタデータ追加
    plan["_generated_at"] = datetime.now().isoformat()
    plan["_idea"] = idea
    plan["_similar_own_posts"] = len(similar_data["own_posts"])
    plan["_similar_trend_posts"] = len(similar_data["trend_posts"])

    # 保存
    _save_plan(plan)

    print("構成案の生成が完了しました！")
    return plan


def _save_plan(plan: Dict[str, Any]):
    """生成された構成案を保存"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # JSON保存
    json_path = DATA_DIR / f"ai_plan_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(plan, f, ensure_ascii=False, indent=2)

    # 最新版
    latest_path = DATA_DIR / "ai_plan_latest.json"
    with open(latest_path, "w", encoding="utf-8") as f:
        json.dump(plan, f, ensure_ascii=False, indent=2)

    # 履歴に追加
    history_path = DATA_DIR / "ai_plan_history.json"
    history = []
    if history_path.exists():
        with open(history_path, "r", encoding="utf-8") as f:
            try:
                history = json.load(f)
            except json.JSONDecodeError:
                history = []

    history.append(plan)
    with open(history_path, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

    print(f"構成案保存: {json_path}")


def export_plan_to_excel(plan: Dict[str, Any], output_path: Optional[str] = None) -> str:
    """構成案をExcelファイルとしてエクスポート（構成案テンプレート形式）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "構成案"

    # スタイル定義
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    content_font = Font(size=10)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    scene_header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # ── ヘッダーセクション ──
    headers = [
        ("テーマ", plan.get("title", "")),
        ("インフルエンサー名", "Kou (@k_slf_imp)"),
        ("動画尺・音声", f"動画尺: {plan.get('duration', '')}\n音声: {plan.get('audio', '')}"),
        ("衣装やロケーション", plan.get("outfit_location", "")),
        ("音源", plan.get("music", "")),
        ("概要欄内容（キャプション、ハッシュタグ）", plan.get("caption", "")),
        ("推奨投稿時間", plan.get("posting_time", "")),
        ("フォーマット", plan.get("format", "")),
    ]

    for i, (label, value) in enumerate(headers, 1):
        cell_label = ws.cell(row=i, column=2, value=label)
        cell_label.font = header_font
        cell_label.fill = header_fill
        cell_label.border = thin_border
        cell_label.alignment = wrap_alignment

        cell_value = ws.cell(row=i, column=3, value=str(value))
        cell_value.font = content_font
        cell_value.border = thin_border
        cell_value.alignment = wrap_alignment
        # 値セルを結合
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=9)

    # 空行
    current_row = len(headers) + 2

    # ── タイトル行 ──
    ws.cell(row=current_row, column=2, value="投稿動画 構成案").font = title_font
    current_row += 1

    # ── シーンヘッダー ──
    scene_headers = [
        "No.", "尺（秒数）", "シーン", "ロケーション",
        "想定着用衣装", "想定時刻", "映像・動作",
        "想定テロップ（文字入れ）", "想定セリフ（音声）",
    ]

    for j, header in enumerate(scene_headers):
        cell = ws.cell(row=current_row, column=j + 1, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = scene_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    current_row += 1

    # ── シーンデータ ──
    scenes = plan.get("scenes", [])
    for scene in scenes:
        row_data = [
            scene.get("scene_number", ""),
            scene.get("duration", ""),
            scene.get("scene_description", ""),
            scene.get("location", ""),
            scene.get("outfit", ""),
            scene.get("time_of_day", ""),
            scene.get("video_action", ""),
            scene.get("text_overlay", ""),
            scene.get("narration", ""),
        ]
        for j, value in enumerate(row_data):
            cell = ws.cell(row=current_row, column=j + 1, value=str(value))
            cell.font = content_font
            cell.border = thin_border
            cell.alignment = wrap_alignment

        ws.row_dimensions[current_row].height = 80
        current_row += 1

    # 空行
    current_row += 1

    # ── ポイント・注意事項 ──
    ws.cell(row=current_row, column=2, value="ポイント・注意事項").font = header_font
    current_row += 1
    tips_cell = ws.cell(row=current_row, column=2, value=plan.get("tips", ""))
    tips_cell.font = content_font
    tips_cell.alignment = wrap_alignment
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
    ws.row_dimensions[current_row].height = 60
    current_row += 1

    # ── 参考分析 ──
    ws.cell(row=current_row, column=2, value="参考データからの知見").font = header_font
    current_row += 1
    ref_cell = ws.cell(row=current_row, column=2, value=plan.get("reference_analysis", ""))
    ref_cell.font = content_font
    ref_cell.alignment = wrap_alignment
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
    ws.row_dimensions[current_row].height = 60

    # 列幅設定
    col_widths = [5, 12, 25, 18, 18, 12, 35, 25, 25]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # ヘッダー行の高さ
    for i in range(1, len(headers) + 1):
        ws.row_dimensions[i].height = max(30, 15 * str(headers[i-1][1]).count("\n") + 30)

    # 保存
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(DATA_DIR / f"構成案_{timestamp}.xlsx")

    wb.save(output_path)
    print(f"Excel保存: {output_path}")
    return output_path


def load_latest_plan() -> Optional[Dict[str, Any]]:
    """最新のAI構成案を読み込み"""
    path = DATA_DIR / "ai_plan_latest.json"
    if not path.exists():
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_plan_history() -> List[Dict[str, Any]]:
    """AI構成案の履歴を読み込み"""
    path = DATA_DIR / "ai_plan_history.json"
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return []


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        idea = " ".join(sys.argv[1:])
    else:
        idea = input("コンテンツアイデアを入力してください: ")

    plan = generate_ai_plan(idea)
    if plan:
        print("\n" + "=" * 50)
        print(f"タイトル: {plan.get('title', '')}")
        print(f"フォーマット: {plan.get('format', '')}")
        print(f"動画尺: {plan.get('duration', '')}")
        print(f"シーン数: {len(plan.get('scenes', []))}")
        print("=" * 50)

        export = input("\nExcelにエクスポートしますか？ (y/n): ")
        if export.lower() == "y":
            path = export_plan_to_excel(plan)
            print(f"エクスポート完了: {path}")
